import io
import urllib
import uuid

from django.conf import settings

import boto3
from celery import shared_task
from celery.utils.log import get_task_logger
import openpyxl

from user.models import User
from catalog.models import Category, Genre
from catalog.utils import slugify

logger = get_task_logger(__name__)


@shared_task(name='core.catalog.bulk.upload.genre.category')
def bulk_upload_category_genre(email: str, file_url: str, username: str, obj_type: str) -> object:
    try:
        user = User.objects.get(username__exact=username)
        file = urllib.request.urlopen(file_url).read()
        work_book = openpyxl.load_workbook(io.BytesIO(file))
        work_sheet = work_book[work_book.sheetnames[0]]
        valid_data = []
        invalid_data = []

        for row in work_sheet.iter_rows():
            if len(row) < 1 or not row[0].value:
                print(len(row), row[0].value)
                break
            if obj_type == 'category':
                slug = slugify(row[0].value)
                try:
                    Category.objects.get(slug__exact=slug)
                    invalid_data.append([row[0].value, 'category already exists'])
                    logger.info('category exists')
                    continue
                except Category.DoesNotExist:
                    category = Category(name=row[0].value, slug=slug, created_by=user)
                    category.save()
                    valid_data.append([row[0].value, 'successfully created'])
                    logger.info('category created')
                except Exception as err:
                    invalid_data.append([row[0].value, f'cannot create category {err}'])
                    logger.info('category cannot be created')
            else:
                slug = slugify(row[0].value)
                try:
                    Genre.objects.get(slug__exact=slug)
                    invalid_data.append([row[0].value, 'genre already exists'])
                    logger.info('genre exists')
                    continue
                except Genre.DoesNotExist:
                    genre = Genre(name=row[0].value, slug=slug, created_by=user)
                    genre.save()
                    valid_data.append([row[0].value, 'successfully created'])
                    logger.info('genre created')
                except Exception as err:
                    invalid_data.append([row[0].value, f'cannot create genre {err}'])
                    logger.info('genre not created')
            logger.info('database insert done!!')
        resp_work_book = openpyxl.Workbook()
        valid_work_book = resp_work_book.active
        valid_work_book.title = 'Completed'
        valid_work_book.column_dimensions['A'].width = 20
        valid_work_book.append(["Name", "Message"])
        for row_data in valid_data:
            valid_work_book.append(row_data)

        if invalid_data:
            invalid_workbook = resp_work_book.create_sheet('Failed')
            invalid_workbook.column_dimensions['A'].width = 20
            invalid_workbook.append(["Name", "Message"])
            for row_data in invalid_data:
                invalid_workbook.append(row_data)
        output = io.BytesIO()
        resp_work_book.save(output)
        s3_client = boto3.client('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
                                 aws_secret_access_key=settings.SECRET_ACCESS_KEY,
                                 region_name='ap-southeast-1')
        filename = f'{uuid.uuid4().hex}.xlsx'
        s3_client.put_object(
            Body=output.getvalue(),
            Bucket=settings.S3_BUCKET,
            Key=filename,
        )
        host = f'https://{settings.S3_BUCKET}.s3.amazonaws.com'
        logger.info(f'{host}/{filename}')
    except Exception as err:
        logger.error(err)
    return
